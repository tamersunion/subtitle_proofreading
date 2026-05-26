#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl

"""
Instruction：

现在需要你写一个程序将一个xlsx（glossary.xlsx）转换为xml

需求如下

1. 整个XML由<Glossary></Glossary>包裹
2. xlsx有3个sheet，他们分别为"数码兽及其技能"、"其他专有名词"、"标题"
3. 第一行为均表头
4. 对于"数码兽及其技能"每一行进行如下处理，各行处理后的xml以换行隔开
    <Digimon JapaneseName="${第A列}" ChineseName="${第B列}">
        <Level DigimonLevel="${第C列}">
        <Skills>
            <Skill JapaneseName="${第D列}" ChineseName="${第E列}"/>
            <Skill JapaneseName="${第F列}" ChineseName="${第G列}"/>
            <Skill JapaneseName="${第H列}" ChineseName="${第I列}"/>
            <Skill JapaneseName="${第J列}" ChineseName="${第K列}"/>
            <Skill JapaneseName="${第L列}" ChineseName="${第M列}"/>
        </Skills>
    </Digimon>
5. 对于"其他专有名词"的每一行进行如下处理，各行处理后的xml以换行隔开
    <Character JapaneseName="${第A列}" ChineseName="${第B列}"/>
    <Entity JapaneseName="${第C列}" ChineseName="${第D列}"/>
    <Term JapaneseName="${第E列}" ChineseName="${第F列}"/>
    <Phrase JapaneseName="${第G列}" ChineseName="${第H列}"/>
6. 对于"标题"的每一行进行处理，各行处理后的xml以换行隔开
    <Title JapaneseName="${第A列}" ChineseName="${第B列}"/>
7. 4~6的内容以换行和<!--Split-->隔开。所有节点里面，如果JapaneseName和ChineseName任意一个为空（判断条件为 len(s.strip())==0），那么就不记录该节点
8. 保存到glossary.xml里面

"""

def safe_str(val):
    if val is None:
        return ""
    return str(val)

def is_valid(s1, s2):
    return len(safe_str(s1).strip()) > 0 and len(safe_str(s2).strip()) > 0

def convert_to_xml(xlsx_path: str = r"C:\Users\kaede\Downloads\驯兽师联盟译名表.xlsx", output_path: str = "glossary.xml"):
    import warnings
    warnings.simplefilter("ignore")

    import openpyxl.styles.stylesheet
    import openpyxl.reader.excel
    orig_apply_stylesheet = openpyxl.styles.stylesheet.apply_stylesheet

    def safe_apply_stylesheet(archive, wb_obj):
        try:
            orig_apply_stylesheet(archive, wb_obj)
        except Exception as err:
            print(f"Warning: Stylesheet ignored due to error: {err}")

    openpyxl.styles.stylesheet.apply_stylesheet = safe_apply_stylesheet
    if hasattr(openpyxl.reader.excel, 'apply_stylesheet'):
        openpyxl.reader.excel.apply_stylesheet = safe_apply_stylesheet

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    sections = []

    # 4. 数码兽及其技能
    if "数码兽及其技能" in wb.sheetnames:
        sheet = wb["数码兽及其技能"]
        lines = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            r = [safe_str(x) for x in row]
            while len(r) < 13:
                r.append("")

            if not is_valid(r[0], r[1]):
                continue

            digimon_xml = f'    <Digimon JapaneseName="{r[0]}" ChineseName="{r[1]}">\n'
            digimon_xml += f'        <Level DigimonLevel="{r[2]}"/>\n'
            digimon_xml += '        <Skills>\n'

            if is_valid(r[3], r[4]):
                digimon_xml += f'            <Skill JapaneseName="{r[3]}" ChineseName="{r[4]}"/>\n'
            if is_valid(r[5], r[6]):
                digimon_xml += f'            <Skill JapaneseName="{r[5]}" ChineseName="{r[6]}"/>\n'
            if is_valid(r[7], r[8]):
                digimon_xml += f'            <Skill JapaneseName="{r[7]}" ChineseName="{r[8]}"/>\n'
            if is_valid(r[9], r[10]):
                digimon_xml += f'            <Skill JapaneseName="{r[9]}" ChineseName="{r[10]}"/>\n'
            if is_valid(r[11], r[12]):
                digimon_xml += f'            <Skill JapaneseName="{r[11]}" ChineseName="{r[12]}"/>\n'

            digimon_xml += '        </Skills>\n'
            digimon_xml += '    </Digimon>'
            lines.append(digimon_xml)
        if lines:
            sections.append("\n".join(lines))

    # 5. 其他专有名词（A/B角色, C/D地点・组织・设施, E/F其他专有名词, G/H固定翻译）
    # 按列分组解析，同类数据集中输出，避免按行交错
    if "其他专有名词" in wb.sheetnames:
        sheet = wb["其他专有名词"]
        characters = []
        entities = []
        terms = []
        phrases = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            r = [safe_str(x) for x in row]
            while len(r) < 8:
                r.append("")

            if is_valid(r[0], r[1]):
                characters.append(f'    <Character JapaneseName="{r[0]}" ChineseName="{r[1]}"/>')
            if is_valid(r[2], r[3]):
                entities.append(f'    <Entity JapaneseName="{r[2]}" ChineseName="{r[3]}"/>')
            if is_valid(r[4], r[5]):
                terms.append(f'    <Term JapaneseName="{r[4]}" ChineseName="{r[5]}"/>')
            if is_valid(r[6], r[7]):
                phrases.append(f'    <Phrase JapaneseName="{r[6]}" ChineseName="{r[7]}"/>')
        lines = characters + entities + terms + phrases
        if lines:
            sections.append("\n".join(lines))

    # 6. 标题
    if "标题" in wb.sheetnames:
        sheet = wb["标题"]
        lines = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            r = [safe_str(x) for x in row]
            while len(r) < 2:
                r.append("")
            if is_valid(r[0], r[1]):
                lines.append(f'    <Title JapaneseName="{r[0]}" ChineseName="{r[1]}"/>')
        if lines:
            sections.append("\n".join(lines))

    # 7. 4~6的内容以换行和<!--Split-->隔开
    inner_xml = "\n<!--Split-->\n".join(sections)

    # 1. 整个XML由<Glossary></Glossary>包裹
    final_xml = f"<Glossary>\n{inner_xml}\n</Glossary>"

    # 8. 保存到glossary.xml里面
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(final_xml)

if __name__ == "__main__":
    convert_to_xml()
